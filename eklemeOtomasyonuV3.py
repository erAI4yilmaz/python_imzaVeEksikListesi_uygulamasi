import os
import sys
import fitz  # PyMuPDF
import customtkinter as ctk
from tkinter import filedialog, messagebox
from PIL import Image
import openpyxl  
import xlrd
import shutil    
import threading 
from openpyxl.utils.cell import coordinate_to_tuple 

# -------------------------------------------------
# GÖRÜNÜM AYARLARI 
# -------------------------------------------------
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# -------------------------------------------------
# EXE / PY ÇALIŞMA DİZİNİ 
# -------------------------------------------------
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)  # ✅ DOĞRU - exe'nin bulunduğu klasör
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TASLAK_ADI = "taslak.xlsx" 

# -------------------------------------------------
# YETKİLİLER
# -------------------------------------------------
YETKILILER = {
    "Çağrı UYGUN": "cagri_uygun.png",
    "Osman ÖZEN": "osman_ozen.png",
    "Oğuzhan MADEN": "oguzhan_maden.png",
    "Kaan ÜRERLER": "kaan_urerler.png",
    "Osman Asilkan YAVUZ": "asilkan_yavuz.png",
    "Doruk TUNÇ": "doruk_tunc.png",
    "Hamdi EROL": "hamdi_erol.png"
}

class ModernSignatureApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Kurumsal İşlem Paneli")
        self.geometry("850x650") 
        self.resizable(False, False)
        
        self.configure(fg_color=("#e2e6e9", "#131314")) 
        
        self.secilen_pdf_yollari = [] 
        self.secilen_excel_yollari = [] 

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # -------------------------------------------------
        # SOL MENÜ (SIDEBAR)
        # -------------------------------------------------
        self.sidebar_frame = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=("#d5dbdf", "#1e1f20"))
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1) 

        logo_yolu = os.path.join(BASE_DIR, "logo.png") 
        if os.path.exists(logo_yolu):
            logo_img = ctk.CTkImage(light_image=Image.open(logo_yolu),
                                    dark_image=Image.open(logo_yolu),
                                    size=(150, 150))
            self.logo_resim_label = ctk.CTkLabel(self.sidebar_frame, text="", image=logo_img)
            self.logo_resim_label.grid(row=0, column=0, pady=(30, 0))
        
        self.logo_label = ctk.CTkLabel(
            self.sidebar_frame, text="MAKRO MÜHENDİSLİK", font=ctk.CTkFont(family="Bahnschrift", size=15, weight="bold"), 
            text_color=("#ffffff", "#041e49"), fg_color=("#516375", "#a8c7fa"), corner_radius=8, padx=15, pady=8                  
        )
        self.logo_label.grid(row=1, column=0, padx=20, pady=(15, 25))

        self.yetkililer_label = ctk.CTkLabel(self.sidebar_frame, text="Kayıtlı Yetkililer:", font=ctk.CTkFont(size=13, weight="bold"), text_color=("#2c3e50", "#e3e3e3"))
        self.yetkililer_label.grid(row=2, column=0, padx=20, pady=(10, 5), sticky="w")

        yetkili_text = "\n".join([f"• {isim}" for isim in YETKILILER.keys()])
        self.liste_label = ctk.CTkLabel(self.sidebar_frame, text=yetkili_text, justify="left", font=ctk.CTkFont(size=12), text_color=("#596775", "#c4c7c5"))
        self.liste_label.grid(row=3, column=0, padx=25, pady=0, sticky="nw")

        self.tema_secici = ctk.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark"], command=self.tema_degistir, 
                                             fg_color=("#bdc3c7", "#333538"), text_color=("#2c3e50", "#e3e3e3"), 
                                             button_color=("#95a5a6", "#444746"), button_hover_color=("#7f8c8d", "#5f6368"))
        self.tema_secici.grid(row=5, column=0, padx=20, pady=20)

        # -------------------------------------------------
        # SAĞ PANEL (ANA EKRAN)
        # -------------------------------------------------
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, padx=35, pady=20, sticky="nsew")

        self.baslik = ctk.CTkLabel(self.main_frame, text="📄 Kurumsal İşlem Paneli", font=ctk.CTkFont(size=24, weight="bold"), text_color=("#2c3e50", "#e3e3e3"))
        self.baslik.pack(anchor="w", pady=(0, 10))

        # --- SEÇİM BÖLÜMÜ ---
        self.card_frame = ctk.CTkFrame(self.main_frame, fg_color=("#edf1f4", "#1e1f20"), border_width=1, border_color=("#bdc3c7", "#444746"), corner_radius=12)
        self.card_frame.pack(fill="x", pady=5, ipady=10)
        self.card_frame.grid_columnconfigure(1, weight=1)

        self.btn_pdf = ctk.CTkButton(self.card_frame, text="📂 PDF SEÇ", command=self.pdf_sec, width=170, height=45, font=ctk.CTkFont(weight="bold"), 
                                     fg_color=("#edf1f4", "#1e1f20"), text_color=("#516375", "#a8c7fa"), 
                                     border_width=1, border_color=("#95a5a6", "#8e918f"), hover_color=("#d5dbdf", "#3c4043"))
        self.btn_pdf.grid(row=0, column=0, padx=20, pady=(15, 5), sticky="w")

        self.secilen_label = ctk.CTkLabel(self.card_frame, text="Henüz PDF seçilmedi", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=13))
        self.secilen_label.grid(row=0, column=1, padx=10, pady=(15, 5), sticky="w")

        self.btn_excel_sec = ctk.CTkButton(self.card_frame, text="📊 EKSİK LİSTESİ SEÇ", command=self.excel_sec, width=170, height=45, font=ctk.CTkFont(weight="bold"), 
                                     fg_color=("#edf1f4", "#1e1f20"), text_color=("#516375", "#a8c7fa"), 
                                     border_width=1, border_color=("#95a5a6", "#8e918f"), hover_color=("#d5dbdf", "#3c4043"))
        self.btn_excel_sec.grid(row=1, column=0, padx=20, pady=(5, 15), sticky="w")

        self.excel_secilen_label = ctk.CTkLabel(self.card_frame, text="Henüz Excel seçilmedi", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=13))
        self.excel_secilen_label.grid(row=1, column=1, padx=10, pady=(5, 15), sticky="w")

        # --- BİLGİ METNİ ---
        self.bilgi_metni = ctk.CTkLabel(self.main_frame, text="\nV2 Yapılan Güncellemeler: \n- Arayüz tasarımı değiştirildi. \n- Birden fazla PDF seçilebilir hale getirildi. \n- Yetkili listesi Eklendi. \n- Ekipman eksik listesi (Toplu Excel okuma) modülü eklendi."
        , justify="left", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=13))
        self.bilgi_metni.pack(anchor="w", pady=(10, 20))

        # --- İŞLEM BUTONLARI ---
        self.btn_imza = ctk.CTkButton(self.main_frame, text="✍️ PDF'LERİ İMZALA", command=self.imzala, width=250, height=45, 
                                      fg_color=("#516375", "#a8c7fa"), hover_color=("#3e4c5a", "#82a1ea"), 
                                      text_color=("#ffffff", "#041e49"), font=ctk.CTkFont(size=15, weight="bold"))
        self.btn_imza.pack(pady=(5, 10))

        self.btn_excel_olustur = ctk.CTkButton(self.main_frame, text="⚙️ EKSİK LİSTESİ OLUŞTUR", command=self.excel_olustur, width=250, height=45, 
                                      fg_color=("#516375", "#a8c7fa"), hover_color=("#3e4c5a", "#82a1ea"), 
                                      text_color=("#ffffff", "#041e49"), font=ctk.CTkFont(size=15, weight="bold"))
        self.btn_excel_olustur.pack(pady=(5, 5))

        # --- GİZLİ YÜKLEME ÇUBUĞU VE DURUM YAZISI ---
        self.progress_bar = ctk.CTkProgressBar(self.main_frame, width=250, fg_color=("#d5dbdf", "#333538"), progress_color=("#2ecc71", "#27ae60"))
        self.progress_bar.set(0)
        
        self.durum_label = ctk.CTkLabel(self.main_frame, text="", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=12))

    # =========================================================================
    # ARAYÜZ SIFIRLAMA (AYRI AYRI GÜNCELLENDİ)
    # =========================================================================
    def pdf_arayuzunu_sifirla(self):
        # Sadece PDF listesini ve etiketini sıfırla
        self.secilen_pdf_yollari = []
        self.secilen_label.configure(text="Henüz PDF seçilmedi", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=13, weight="normal"))
        
        # Barı gizle ve butonları aç (ortak işlem)
        self._ortak_kilitleri_ac()

    def excel_arayuzunu_sifirla(self):
        # Sadece Excel listesini ve etiketini sıfırla
        self.secilen_excel_yollari = []
        self.excel_secilen_label.configure(text="Henüz Excel seçilmedi", text_color=("#596775", "#c4c7c5"), font=ctk.CTkFont(size=13, weight="normal"))
        
        # Barı gizle ve butonları aç (ortak işlem)
        self._ortak_kilitleri_ac()

    def _ortak_kilitleri_ac(self):
        # Yükleme barını ve durum yazısını gizle
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()
        self.durum_label.configure(text="")
        self.durum_label.pack_forget()
        
        # Tüm butonların kilitlerini aç
        self.btn_imza.configure(state="normal")
        self.btn_excel_olustur.configure(state="normal")
        self.btn_pdf.configure(state="normal")
        self.btn_excel_sec.configure(state="normal")

    # --- TEMA VE DOSYA SEÇİMİ ---
    def tema_degistir(self, yeni_tema: str):
        ctk.set_appearance_mode(yeni_tema)

    def pdf_sec(self):
        dosyalar = filedialog.askopenfilenames(title="İmzalanacak PDF'leri Seç", filetypes=[("PDF Dosyaları", "*.pdf")])
        if dosyalar:
            self.secilen_pdf_yollari = list(dosyalar)
            dosya_sayisi = len(self.secilen_pdf_yollari)
            metin = os.path.basename(self.secilen_pdf_yollari[0]) if dosya_sayisi == 1 else f"Toplam {dosya_sayisi} adet PDF seçildi."
            self.secilen_label.configure(text=metin, text_color=("#3e4c5a", "#a8c7fa"), font=ctk.CTkFont(size=13, weight="bold"))

    def excel_sec(self):
        dosyalar = filedialog.askopenfilenames(title="Kaynak Excel'leri Seç", filetypes=[("Excel Dosyaları", "*.xlsx *.xlsm *.xls"),("Tüm Dosyalar","*.*")])
        if dosyalar:
            self.secilen_excel_yollari = list(dosyalar)
            dosya_sayisi = len(self.secilen_excel_yollari)
            metin = os.path.basename(self.secilen_excel_yollari[0]) if dosya_sayisi == 1 else f"Toplam {dosya_sayisi} adet Excel seçildi."
            self.excel_secilen_label.configure(text=metin, text_color=("#3e4c5a", "#a8c7fa"), font=ctk.CTkFont(size=13, weight="bold"))

    # =========================================================================
    # 1) PDF İMZALAMA MOTORU
    # =========================================================================
    def imzala(self):
        if not self.secilen_pdf_yollari:
            messagebox.showwarning("Uyarı", "ÖNCE PDF SEÇMELİSİNİZ.")
            return

        # Kullanıcı işlem yaparken her yeri kilitle
        self.btn_imza.configure(state="disabled")
        self.btn_excel_olustur.configure(state="disabled")
        self.btn_pdf.configure(state="disabled")
        self.btn_excel_sec.configure(state="disabled")
        
        self.progress_bar.pack(pady=(10, 0))
        self.durum_label.pack(pady=(5, 0))
        self.durum_label.configure(text="Sistem hazırlanıyor...")
        self.progress_bar.set(0.0)

        threading.Thread(target=self._pdf_islem_arkaplan, daemon=True).start()
    
        
    
    
    
    def xls_to_xlsx(self, xls_yolu):
    
        import tempfile
        xls_wb = xlrd.open_workbook(xls_yolu)
        yeni_wb = openpyxl.Workbook()
        yeni_wb.remove(yeni_wb.active)

        for sayfa_adi in xls_wb.sheet_names():
            xls_ws = xls_wb.sheet_by_name(sayfa_adi)
            yeni_ws = yeni_wb.create_sheet(title=sayfa_adi)
            for row in range(xls_ws.nrows):
                for col in range(xls_ws.ncols):
                    yeni_ws.cell(row=row+1, column=col+1, value=xls_ws.cell_value(row, col))

        gecici = tempfile.mktemp(suffix=".xlsx")
        yeni_wb.save(gecici)
        return gecici
    
    
    
    
    
    
    
    
    
    
    
    
    
    def _pdf_islem_arkaplan(self):
        basarili_sayisi = 0
        hatali_dosyalar = []
        toplam_pdf = len(self.secilen_pdf_yollari)

        try:
            for index, pdf_yolu in enumerate(self.secilen_pdf_yollari):
                self.durum_label.configure(text=f"İmzalanıyor ({index+1}/{toplam_pdf}): {os.path.basename(pdf_yolu)}")
                self.progress_bar.set(index / toplam_pdf)

                try:
                    doc = fitz.open(pdf_yolu)
                    imza_basildi = False
                    sayfa_sayisi = len(doc)

                    for s_index, sayfa in enumerate(doc):
                        ilerleme_orani = (index / toplam_pdf) + ((s_index / sayfa_sayisi) * (1 / toplam_pdf))
                        self.progress_bar.set(ilerleme_orani)

                        sayfa_metni = sayfa.get_text()
                        bulunan_resim = None
                        
                        for isim, resim in YETKILILER.items():
                            if isim.casefold() in sayfa_metni.casefold():
                                bulunan_resim = resim
                                break

                        if not bulunan_resim:
                            continue

                        imza_list = sayfa.search_for("İMZA")
                        muhur_list = sayfa.search_for("MÜHÜR")

                        if not imza_list or not muhur_list:
                            continue

                        imza_rect = imza_list[0]
                        muhur_rect = muhur_list[0]
                        png_yolu = os.path.join(BASE_DIR, bulunan_resim)

                        if not os.path.exists(png_yolu):
                            hatali_dosyalar.append(f"{os.path.basename(pdf_yolu)} ({bulunan_resim} bulunamadı)")
                            break

                        hucre_genislik = muhur_rect.x0 - imza_rect.x0
                        hucre_yukseklik = imza_rect.height * 6
                        ust_y = imza_rect.y1 + 5
                        
                        pix = fitz.Pixmap(png_yolu)
                        oran = pix.width / pix.height
                        yeni_genislik = hucre_genislik * 1.9
                        yeni_yukseklik = yeni_genislik / oran

                        if yeni_yukseklik > hucre_yukseklik:
                            yeni_yukseklik = hucre_yukseklik
                            yeni_genislik = yeni_yukseklik * oran

                        merkez_x = (imza_rect.x0 + muhur_rect.x0) / 2
                        hedef = fitz.Rect(merkez_x - yeni_genislik / 2, ust_y, merkez_x + yeni_genislik / 2, ust_y + yeni_yukseklik)

                        sayfa.insert_image(hedef, filename=png_yolu)
                        imza_basildi = True

                    if not imza_basildi:
                        hatali_dosyalar.append(f"{os.path.basename(pdf_yolu)} (Uygun alan bulunamadı)")
                        doc.close()
                        continue

                    cikti_yolu = os.path.join(os.path.dirname(pdf_yolu), "İmzalı-" + os.path.basename(pdf_yolu))
                    doc.save(cikti_yolu)
                    doc.close()
                    basarili_sayisi += 1

                except Exception as e:
                    hatali_dosyalar.append(f"{os.path.basename(pdf_yolu)} (Hata: {str(e)})")

            self.durum_label.configure(text="PDF İmzalama Tamamlandı!")
            self.progress_bar.set(1.0)
            
            ozet = f"{basarili_sayisi} adet PDF imzalandı."
            if hatali_dosyalar:
                ozet += "\n\n⚠️ Eksikler:\n" + "\n".join(hatali_dosyalar)
                messagebox.showwarning("Tamamlandı", ozet)
            else:
                messagebox.showinfo("Başarılı", ozet)

        finally:
            # GÜNCELLENDİ: Sadece PDF tarafını sıfırla
            self.pdf_arayuzunu_sifirla()


    # =========================================================================
    # 2) EXCEL OLUŞTURMA MOTORU 
    # =========================================================================
    def excel_olustur(self):
        if not self.secilen_excel_yollari:
            messagebox.showwarning("Uyarı", "ÖNCE KAYNAK EXCEL SEÇMELİSİNİZ.")
            return

        taslak_yolu = os.path.join(BASE_DIR, TASLAK_ADI)
        if not os.path.exists(taslak_yolu):
            messagebox.showerror("Hata", f"'{TASLAK_ADI}' bulunamadı!\nLütfen taslak dosyasını uygulamanın yanına koyun.")
            return

        self.btn_excel_olustur.configure(state="disabled")
        self.btn_imza.configure(state="disabled")
        self.btn_pdf.configure(state="disabled")
        self.btn_excel_sec.configure(state="disabled")
        
        self.progress_bar.pack(pady=(10, 0))
        self.durum_label.pack(pady=(5, 0))
        self.durum_label.configure(text="Sistem hazırlanıyor...")
        self.progress_bar.set(0.0)

        threading.Thread(target=self._excel_islem_arkaplan, args=(taslak_yolu,), daemon=True).start()

    def hucre_oku(self, sheet_data, koordinat):
        try:
            row, col = coordinate_to_tuple(koordinat)
            if row - 1 < len(sheet_data):
                satir_verisi = sheet_data[row - 1]
                if col - 1 < len(satir_verisi):
                    return satir_verisi[col - 1]
        except:
            pass
        return None

    def _excel_islem_arkaplan(self, taslak_yolu):
        try:
            kayit_dizini = os.path.dirname(self.secilen_excel_yollari[0])
            kayit_yolu = os.path.join(kayit_dizini, "-Eksik Listesi.xlsx")
            shutil.copy(taslak_yolu, kayit_yolu)

            hedef_wb = openpyxl.load_workbook(kayit_yolu)
            hedef_ws = hedef_wb.active 

            satir = 3 
            toplam_dosya = len(self.secilen_excel_yollari)

            for d_index, excel_yolu in enumerate(self.secilen_excel_yollari):
                dosya_adi = os.path.basename(excel_yolu)
                
                
                gercek_yol = self.xls_to_xlsx(excel_yolu) if excel_yolu.endswith(".xls") else excel_yolu
                kaynak_wb = openpyxl.load_workbook(gercek_yol, data_only=True, read_only=True)
                sayfalar = kaynak_wb.sheetnames
                toplam_sayfa = len(sayfalar)

                for s_index, sayfa_adi in enumerate(sayfalar):
                    self.durum_label.configure(text=f"Dosya ({d_index+1}/{toplam_dosya}) | Sayfa Taranıyor: {sayfa_adi}")
                    ilerleme_orani = (d_index / toplam_dosya) + ((s_index / toplam_sayfa) * (1 / toplam_dosya))
                    self.progress_bar.set(ilerleme_orani)

                    isim = sayfa_adi.lower()
                    
                    gerekli_sayfa = any(k in isim for k in [
"kompresör","aed","araç taşıyıcı","caraskal","cephe asansörü","iskele",
"çektirme","ytt","forklift","yoyo","araç kaldırma lifti","istif mak.",
"kaldırma ekipmanı","kaldırma makinası","kaldırma tablası","kanca altı",
"kapı","birim konveyör","yığın konveyör","kriko","kule vinç","manipülatör",
"mapa","merdiven","mobil vinç","rampa","sapan","transpalet","vinç",
"yürüyen merdiven",

"basınçlı hava tankı","basınçlı kap","bk emniyet cihazı","yıkama mak.",
"boyama kazanı","boyler","buhar jen.","buhar jeneratörü","buhar kaz.",
"buhar kazanı","hid.tankı","kara tankeri","tehlikeli sıvı","kızgın su",
"kızgın yağ","motopomp","san.gaz.tankeri","sıcak su kaz.","sıv.gaz tank.",
"ütü kaz.","yemek pişirme kazanı",

"bükme","chiller","cnc kesme","diş çekme","eğlence","enjeksiyon",
"erezyon","ısı eşanjörü","freze","demir kesme","kaynak","makas",
"matkap","planya","pres","taşlama","torna","zımpara"
])
                    
                    if gerekli_sayfa:
                        ws = kaynak_wb[sayfa_adi]
                        sheet_data = list(ws.values) 



                    #===========KALDIRMA EKİPMANLARI======================


                        if "kompresör" in isim:
                            data = [self.hucre_oku(sheet_data, "V21"), self.hucre_oku(sheet_data, "V20"), self.hucre_oku(sheet_data, "BO17"), self.hucre_oku(sheet_data, "V17"), self.hucre_oku(sheet_data, "DB17")]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 54))
                            satir += 1

                        elif "aed" in isim:
                            data = [
                                "AED",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "AE27"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R17"),     # E → Markası
                                self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 64))
                            satir += 1
                        
                        
                        elif "araç taşıyıcı" in isim:
                            data = [
                                "Araç Taşıyıcı",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "DC17"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "R18"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R15"),     # E → Markası
                                self.hucre_oku(sheet_data, "BP15"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(44, 62))
                            satir += 1
                        
                        elif "caraskal" in isim:
                            data = [
                                "Caraskal",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "BO23"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "R19"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R17"),     # E → Markası
                                self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 56))
                            satir += 1
                            
                        elif "cephe asansörü" in isim:
                            data = [
                                "Cephe Asansörü",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "R21"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R17"),     # E → Markası
                                self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 56))
                            satir += 1
                        
                        
                        elif "iskele" in isim:
                            data = [
                                "İskele",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "DB22"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R17"),     # E → Markası
                                self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 56))
                            satir += 1
                        
                        elif "çektirme" in isim:
                            data = [
                                "Çektirme",                       # B → sabit metin
                                self.hucre_oku(sheet_data, "R23"),     # C → İşletmedeki yeri
                                self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                                self.hucre_oku(sheet_data, "R16"),     # E → Markası
                                self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                    ]
                            self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 57))
                            satir += 1
                            
                            
                        elif "ytt" in isim:
                             data = [
                                 "Eğimli Yük Taşıma Tertibatı",                       # B → sabit metin
                                 self.hucre_oku(sheet_data, "R21"),     # C → İşletmedeki yeri
                                 self.hucre_oku(sheet_data, "R18"),     # D → Seri no
                                 self.hucre_oku(sheet_data, "R16"),     # E → Markası
                                 self.hucre_oku(sheet_data, "V23"),     # F → Kapasitesi
                                     ]
                             self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(55, 61))
                             satir += 1
                        
                        elif "forklift" in isim:
                           data = [
                               "Forklift",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(53, 64))
                           satir += 1 
                        
                        
                        
                        elif "yoyo" in isim:
                           data = [
                               "YOYO",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 59))
                           satir += 1 
                        
                        elif "araç kaldırma lifti" in isim:
                           data = [
                               "Araç Kaldırma Lifti",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(65, 66))
                           satir += 1 
                        
                        
                        elif "istif mak." in isim or "istif makinası" in isim:
                           data = [
                               "İstif Makinası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 66))
                           satir += 1 
                        
                        
                        elif "kaldırma ekipmanı" in isim:
                           data = [
                               "Kaldırma İletme Ekipmanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "Z20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(48, 56))
                           satir += 1
                           
                           
                        elif "kaldırma makinası" in isim:
                           data = [
                               "Kaldırma İletme Makinası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "Z19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(60, 63))
                           satir += 1
                        
                        
                        elif "kaldırma tablası" in isim:
                           data = [
                               "Kaldırma Tablası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 60))
                           satir += 1
                        
                        
                        
                        elif "kanca altı" in isim:
                           data = [
                               "Kanca Altı Kaldırma Ekipmanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 57))
                           satir += 1
                        
                        elif "kapı" in isim or "endüstriyel kapı" in isim:
                           data = [
                               "endüstriyel kapı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BO21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R18"),     # E → Markası
                               "-",     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(38, 56))
                           satir += 1
                        
                        elif "birim konveyör" in isim:
                           data = [
                               "Birim Konveyör",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 57))
                           satir += 1
                        
                        elif "yığın konveyör" in isim:
                           data = [
                               "Yığın Konveyör",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(46, 60))
                           satir += 1
                        
                        
                        elif "kriko" in isim:
                           data = [
                               "kriko",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BO19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(45, 58))
                           satir += 1
                        
                        elif "kule vinç" in isim:
                           data = [
                               "Kule Vinç",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R25"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(45, 57))
                           satir += 1
                        
                        
                        elif "manipülatör" in isim:
                           data = [
                               "Manipülatör",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 58))
                           satir += 1
                        
                        elif "mapa" in isim:
                           data = [
                               "Mapa",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BO21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(38, 53))
                           satir += 1
                        
                        
                        
                        elif "merdiven" in isim:
                           data = [
                               "Merdiven",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R23"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO20"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R20"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD20"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(38, 53))
                           satir += 1
                        
                        elif "mobil vinç" in isim:
                           data = [
                               "Mobil Vinç",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "CT25"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO20"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(61, 65))
                           satir += 1
                        
                        elif "taşıt üstü yükleyici" in isim:
                           data = [
                               "taşıt üstü yükleyici",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "CT25"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO20"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(61, 64))
                           satir += 1
                        
                        
                        elif "rampa" in isim:
                           data = [
                               "Rampa",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 60))
                           satir += 1
                        
                        
                        
                        elif "sabit iniş mah." in isim:
                           data = [
                               "Sabit İniş mah.",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R24"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO22"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R21"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD21"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(84, 86))
                           satir += 1
                        
                        
                        
                        elif "sapan" in isim:
                           data = [
                               "Sapan.",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R25"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO22"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R22"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD22"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 57))
                           satir += 1
                        
                        
                        elif "iskele" in isim or "seyyar erişim kulesi" in isim:
                           data = [
                               "Seyyar Erişim Kulesi",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "R23"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO20"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R20"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD20"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(44, 54))
                           satir += 1
                        
                        
                        
                        elif "platform" in isim:
                           data = [
                               "Sütünlu Çaışma Platformu",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BP22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(46, 48))
                           satir += 1
                        
                        elif "transpalet" in isim:
                           data = [
                               "Transpalet",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(48, 62))
                           satir += 1
                        
                        
                        elif "vinç" in isim or "vinc" in isim:
                           data = [
                               "Vinç",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "CT26"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R20"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(53, 56))
                           satir += 1
                        
                        
                        
                    #yükseltilebilir iş platformu vardı -------------
                        
                        
                        
                        elif "yürüyen merdiven" in isim:
                           data = [
                               "Yürüyen Merdiven",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BO16"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R15"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 54))
                           satir += 1
                        
                        
                        
                        
                        ###=============BASINÇLI KAPLAR===========================================
                        
                        
                        
                        
                        elif "basınçlı hava tankı" in isim:
                           data = [
                               "Basınçlı Hava Tankı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(44, 52))
                           satir += 1
                        
                        
                        elif "basınçlı kap" in isim:
                           data = [
                               "Basınçlı Kap",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 55))
                           satir += 1
                        
                        
                        elif "bk emniyet cihazı" in isim:
                           data = [
                               "BK Emniyet Cihazı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(40, 53))
                           satir += 1
                        
                        
                        
                        elif "basınçlı hava tankı" in isim:
                           data = [
                               "basınçlı hava tankı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V18"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V14"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(44, 55))
                           satir += 1
                        
                        
                        elif "yıkama mak." in isim:
                           data = [
                               "Basınçlı Yıkama Makinası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO17"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB21"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 54))
                           satir += 1
                        
                        
                        elif "boyama kazanı" in isim:
                           data = [
                               "Boyama Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(45, 55))
                           satir += 1
                        
                        
                        
                        
                        elif "boyler" in isim:
                           data = [
                               "Boyler",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(41, 57))
                           satir += 1
                        
                        
                        elif "buhar jen." in isim or "buhar jeneratörü" in isim:
                           data = [
                               "Buhar Jeneratörü",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 59))
                           satir += 1
                        
                        elif "buhar kaz." in isim or "buhar kazanı" in isim:
                           data = [
                               "Buhar Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V20"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 58))
                           satir += 1
                        
                        elif "hid.tankı" in isim:
                           data = [
                               "Hidrofor/Genleşme Tankı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V19"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB19"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 54))
                           satir += 1
                        
                        
                        
                        elif "kara tankeri" in isim:
                           data = [
                               "kara tankeri tankı",                       # B → sabit metin
                               "-",     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(60, 63))
                           satir += 1
                        
                        
                        elif "hid.tankı" in isim:
                           data = [
                               "Hidrofor/Genleşme Tankı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V19"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB19"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 54))
                           satir += 1
                        
                        
                        elif "tehlikeli sıvı" in isim:
                           data = [
                               "Kimyasal Depolama Dolabı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V19"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V16"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(47, 55))
                           satir += 1
                        
                        
                        elif "kızgın su" in isim:
                           data = [
                               "Kızgın Su Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V20"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB18"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(53, 57))
                           satir += 1
                        
                        elif "kızgın yağ" in isim:
                           data = [
                               "Kızgın Yağ Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 56))
                           satir += 1
                        
                        
                        elif "motopomp" in isim:
                           data = [
                               "Motopomp",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V18"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB19"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(46, 57))
                           satir += 1
                        
                        
                        
                        
                        elif "san.gaz.tankeri" in isim:
                           data = [
                               "Sanayi Gazları Tankeri",                       # B → sabit metin
                               "-",     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V25"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V22"),     # E → Markası
                               self.hucre_oku(sheet_data, "V27"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(60, 62))
                           satir += 1
                        
                        
                        elif "sıcak su kaz." in isim:
                           data = [
                               "Sıcak Su Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V22"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V19"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB19"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(48, 55))
                           satir += 1
                        
                        elif "sıv.gaz tank." in isim:
                           data = [
                               "Sıvı Gaz Tankı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V15"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V13"),     # E → Markası
                               self.hucre_oku(sheet_data, "V17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(58, 60))
                           satir += 1
                        
                        
                        
                        elif "ütü kaz." in isim:
                           data = [
                               "Ütü Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V19"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V17"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO17"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(44, 54))
                           satir += 1
                        
                        
                        elif "yemek pişirme kazanı" in isim:
                           data = [
                               "Yemek Pişirme Kazanı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "V20"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "V18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "V16"),     # E → Markası
                               self.hucre_oku(sheet_data, "DB16"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(45, 54))
                           satir += 1
                        
                        
                        
                        
                        
                        #===============İŞ MAKİNALARI=====================
                        
                        
                        
                        elif "arazöz" in isim:
                           data = [
                               "Arazöz",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BP15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BP13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BP14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 57))
                           satir += 1
                        
                        elif "beton mikseri" in isim:
                           data = [
                               "Beton Mikseri",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(49, 59))
                           satir += 1
                        
                        
                        
                        elif "beton pompası" in isim:
                           data = [
                               "Beton Pompası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "DD14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 60))
                           satir += 1
                        
                        
                        elif "çekici" in isim:
                           data = [
                               "Çekici",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 60))
                           satir += 1
                        
                        
                        elif "dozer" in isim:
                           data = [
                               "Dozer",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BQ15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BQ13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               "-",     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 60))
                           satir += 1
                        
                        
                        elif "asfaltlama" in isim:
                           data = [
                               "Finişer",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BQ15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BQ13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               "-",     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 63))
                           satir += 1
                        
                        
                        elif "fore kazık" in isim:
                           data = [
                               "Fore Kazık",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(64, 66))
                           satir += 1
                        
                        
                        elif "greyder" in isim:
                           data = [
                               "Greyder",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BQ15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BQ13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               "-",     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 61))
                           satir += 1
                        
                        
                        elif "jeneratör" in isim:
                           data = [
                               "Jeneratör",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(42, 61))
                           satir += 1
                        
                        
                        
                        elif "kamyon" in isim:
                           data = [
                               "Kamyon",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(55, 62))
                           satir += 1
                        
                        elif "kazıyıcı yükleyici" in isim:
                           data = [
                               "Kazıyıcı Yükleyici",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 61))
                           satir += 1
                        
                        
                        
                        
                        elif "silobas" in isim:
                           data = [
                               "Silobas",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DD17"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(46, 58))
                           satir += 1
                        
                        
                        elif "skreyper" in isim:
                           data = [
                               "Skreyper",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BQ15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BQ13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               "-",     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 62))
                           satir += 1
                        
                        
                        elif "yükleyici" in isim:
                           data = [
                               "Yükleyici",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "BQ15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BQ13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 62))
                           satir += 1
                        
                        elif "silindir" in isim:
                           data = [
                               "Silindir",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DC15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BP13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 60))
                           satir += 1
                        
                        
                        
                        
                        #=============TEZGAH===========================
                        
                        
                        elif "bükme" in isim:
                           data = [
                               "Bükme Tezgahı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(56, 66))
                           satir += 1
                        
                        
                        elif "chıller" in isim:
                           data = [
                               "Chiller",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DC14"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(56, 66))
                           satir += 1
                        
                        
                        
                        elif "cnc kesme" in isim:
                           data = [
                               "CNC Lazer/Plazma Tezgahı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(60, 66))
                           satir += 1
                        
                        
                        elif "diş çekme" in isim:
                           data = [
                               "Diş Çekme",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(50, 60))
                           satir += 1
                        
                        
                        
                        elif "eğlence" in isim:
                           data = [
                               "Eğlence Makinası",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 59))
                           satir += 1
                        
                        
                        
                        elif "enjeksiyon" in isim:
                           data = [
                               "Enjeksiyon Tezgahı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 62))
                           satir += 1
                        
                        
                        elif "erezyon" in isim:
                           data = [
                               "Erozyon Tezgahı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(54, 62))
                           satir += 1
                        
                        elif "ısı eşanjörü" in isim:
                           data = [
                               "Isı Eşanjörü",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "AF21"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "CP18"),     # D → Seri no
                               self.hucre_oku(sheet_data, "AF18"),     # E → Markası
                               self.hucre_oku(sheet_data, "AF25"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(51, 57))
                           satir += 1
                        
                        
                        elif "freze" in isim:
                           data = [
                               "Freze",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(58, 63))
                           satir += 1
                        
                        
                        
                        elif "demir kesme" in isim:
                           data = [
                               "Demir Kesme",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 65))
                           satir += 1
                        
                        elif "kaynak" in isim:
                           data = [
                               "Kaynak",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO13"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(61, 65))
                           satir += 1
                        
                        elif "makas" in isim:
                           data = [
                               "Giyotin Makas",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(62, 66))
                           satir += 1
                        
                        
                        elif "matkap" in isim:
                           data = [
                               "Delme Matkap",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB15"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "R16"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(57, 63))
                           satir += 1
                        
                        
                        elif "pafta" in isim:
                           data = [
                               "Pafta Tezgahı",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 56))
                           satir += 1
                        
                        
                        
                        elif "planya" in isim:
                           data = [
                               "Planya",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 56))
                           satir += 1
                       
                        elif "pres" in isim:
                           data = [
                               "Pres",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB14"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO14"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(82, 85))
                           satir += 1
                        
                        
                        
                        elif "taşlama" in isim:
                           data = [
                               "Taşlama",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(56, 63))
                           satir += 1
                        
                        
                        elif "torna" in isim:
                           data = [
                               "Torna",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(58, 62))
                           satir += 1
                        
                        
                        
                        elif "zımpara" in isim:
                           data = [
                               "Zımpara",                       # B → sabit metin
                               self.hucre_oku(sheet_data, "DB13"),     # C → İşletmedeki yeri
                               self.hucre_oku(sheet_data, "BO13"),     # D → Seri no
                               self.hucre_oku(sheet_data, "R13"),     # E → Markası
                               self.hucre_oku(sheet_data, "BO15"),     # F → Kapasitesi
                                   ]
                           self._excel_veriyi_yaz_ram(hedef_ws, satir, data, sheet_data, range(52, 63))
                           satir += 1
                        
                        
                        
                        
                        
                        
                        
                        
                        
                        
                        
                        

                kaynak_wb.close() 
                
            hedef_wb.save(kayit_yolu)
            
            self.durum_label.configure(text="Tüm Dosyalar Başarıyla Tamamlandı!")
            self.progress_bar.set(1.0)
            messagebox.showinfo("Başarılı", f"Seçilen tüm dosyalar tarandı ve eksik listesi oluşturuldu!")

        except Exception as e:
            self.durum_label.configure(text="Hata oluştu!")
            messagebox.showerror("Hata", f"İşlem sırasında bir hata oluştu: {str(e)}")
        finally:
            # GÜNCELLENDİ: Sadece Excel tarafını sıfırla
            self.excel_arayuzunu_sifirla()

    def _excel_veriyi_yaz_ram(self, hedef_ws, satir, data, sheet_data, d_araligi):
        for i, deger in enumerate(data, start=2):
            hedef_ws.cell(row=satir, column=i, value=deger)
        
        notlar = [str(self.hucre_oku(sheet_data, f"A{r}")) for r in d_araligi if self.hucre_oku(sheet_data, f"A{r}")]
        final_not = "\n".join(notlar) if notlar else "Uygun"
        
        cell = hedef_ws.cell(row=satir, column=7, value=final_not)
        cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="center")

if __name__ == "__main__":
    app = ModernSignatureApp()
    app.mainloop()